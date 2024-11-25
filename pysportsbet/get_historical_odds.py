import requests
import openpyxl
from datetime import datetime, timedelta

# Configuration constants
SPREADSHEET_FILE = "odds_data.xlsx"  # Path to the output Excel file
SHEET_NAME = "Sheet1"  # Name of the sheet in the Excel file
API_KEY = "YOUR_API_KEY"  # Replace with your API key from The Odds API
SPORT_KEY = "baseball_mlb"  # Sport key for the desired sport
MARKETS = "h2h,spreads,totals"  # Comma-separated list of betting markets
REGIONS = "us"  # Supported regions: us, uk, eu, au
BOOKMAKERS = ""  # Optional: specific bookmakers
ODDS_FORMAT = "american"  # Odds format: 'american' or 'decimal'
FROM_DATE = "2023-09-10T00:00:00Z"  # Start date
TO_DATE = "2023-09-10T12:00:00Z"  # End date
INTERVAL_MINS = 60  # Interval between snapshots in minutes

HEADERS = [
    "timestamp", "id", "commence_time", "bookmaker", "last_update",
    "home_team", "away_team", "market", "label_1", "odd_1", "point_1",
    "label_2", "odd_2", "point_2", "odd_draw"
]


def fetch_odds(api_key, sport_key, regions, bookmakers, markets, odds_format, timestamp):
    """
    Fetch historical odds data for a specific sport and timestamp.

    Args:
        api_key (str): The Odds API key.
        sport_key (str): Sport key (e.g., 'baseball_mlb').
        regions (str): Comma-separated list of regions (e.g., 'us').
        bookmakers (str): Comma-separated list of bookmakers.
        markets (str): Comma-separated list of markets.
        odds_format (str): Format of the odds ('american' or 'decimal').
        timestamp (str): Date and time of the snapshot (ISO 8601 format).

    Returns:
        dict: A dictionary containing metadata and response content.
    """
    bookmakers_param = f"bookmakers={bookmakers}" if bookmakers else f"regions={regions}"
    url = f"https://api.the-odds-api.com/v4/historical/sports/{sport_key}/odds"
    params = {
        "apiKey": api_key,
        bookmakers_param: bookmakers_param,
        "markets": markets,
        "oddsFormat": odds_format,
        "date": timestamp
    }
    response = requests.get(url, params=params)
    response.raise_for_status()
    return {
        "metaData": format_response_meta_data(response.headers),
        "responseContent": response.json()
    }


def format_event_output(response):
    """
    Format the API response into rows suitable for Excel output.

    Args:
        response (dict): API response containing event data.

    Returns:
        list: A list of rows formatted for output.
    """
    rows = []
    for event in response.get("data", []):
        for bookmaker in event.get("bookmakers", []):
            for market in bookmaker.get("markets", []):
                if market["key"] == "totals":
                    outcome_home = next((o for o in market["outcomes"] if o["name"] == "Over"), {})
                    outcome_away = next((o for o in market["outcomes"] if o["name"] == "Under"), {})
                    outcome_draw = {}
                else:
                    outcome_home = next((o for o in market["outcomes"] if o["name"] == event["home_team"]), {})
                    outcome_away = next((o for o in market["outcomes"] if o["name"] == event["away_team"]), {})
                    outcome_draw = next((o for o in market["outcomes"] if o["name"] == "Draw"), {})

                rows.append([
                    response["timestamp"], event["id"], event["commence_time"],
                    bookmaker["key"], bookmaker["last_update"], event["home_team"],
                    event["away_team"], market["key"], outcome_home.get("name"),
                    outcome_home.get("price"), outcome_home.get("point"),
                    outcome_away.get("name"), outcome_away.get("price"),
                    outcome_away.get("point"), outcome_draw.get("price")
                ])
    return rows


def format_response_meta_data(headers):
    """
    Extract metadata from response headers.

    Args:
        headers (dict): Response headers.

    Returns:
        list: A list of metadata rows.
    """
    return [
        ["Requests Used", headers.get("x-requests-used")],
        ["Requests Remaining", headers.get("x-requests-remaining")]
    ]


def main():
    """
    Main function to fetch historical odds data and save it to an Excel file.

    Writes:
        Excel file specified in SPREADSHEET_FILE.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # Write headers
    ws.append(HEADERS)

    # Parse date strings into datetime objects
    from_datetime = datetime.fromisoformat(FROM_DATE[:-1])
    to_datetime = datetime.fromisoformat(TO_DATE[:-1])
    current_datetime = to_datetime

    # Initialize output row
    output_row = 2

    while current_datetime > from_datetime:
        formatted_date = current_datetime.isoformat() + "Z"
        print(f"Fetching odds for {formatted_date}")

        # Fetch odds data
        data = fetch_odds(API_KEY, SPORT_KEY, REGIONS, BOOKMAKERS, MARKETS, ODDS_FORMAT, formatted_date)

        # Write metadata to Excel
        for i, meta_row in enumerate(data["metaData"]):
            ws.cell(row=i + 1, column=1, value=meta_row[0])
            ws.cell(row=i + 1, column=2, value=meta_row[1])

        # Write odds data to Excel
        formatted_response = format_event_output(data["responseContent"])
        for row in formatted_response:
            ws.append(row)
        output_row += len(formatted_response)

        # Update the current timestamp
        previous_timestamp = data["responseContent"].get("previous_timestamp")
        if not previous_timestamp:
            print("No earlier historical data available.")
            break
        current_datetime = min(
            current_datetime - timedelta(minutes=INTERVAL_MINS),
            datetime.fromisoformat(previous_timestamp[:-1])
        )

    # Save the workbook
    wb.save(SPREADSHEET_FILE)
    print(f"Data saved to {SPREADSHEET_FILE}")


if __name__ == "__main__":
    main()
