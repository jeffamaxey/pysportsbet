import requests
import openpyxl
from datetime import datetime, timedelta

# Configuration constants
SPREADSHEET_FILE = "odds_data.xlsx"  # Path to the output Excel file
SHEET_NAME = "Sheet1"  # Name of the sheet in the Excel file
API_KEY = "YOUR_API_KEY"  # Replace with your API key from The Odds API
SPORT_KEY = "baseball_mlb"  # Sport key for the desired sport
MARKETS = "h2h,spreads,totals"  # Comma-separated list of betting markets
REGIONS = "us"  # Supported regions: us, uk, eu, au
BOOKMAKERS = ""  # Optional: specific bookmakers
ODDS_FORMAT = "american"  # Odds format: 'american' or 'decimal'
FROM_DATE = "2023-09-10T00:00:00Z"  # Start date
TO_DATE = "2023-09-10T12:00:00Z"  # End date
INTERVAL_MINS = 60  # Interval in minutes

HEADERS = [
    "timestamp", "id", "commence_time", "bookmaker", "home_team", "away_team",
    "market", "last_update", "label", "description", "price", "point"
]


def fetch_events(api_key, sport_key, timestamp):
    """
    Fetch historical events from The Odds API for a given sport and timestamp.
    """
    url = f"https://api.the-odds-api.com/v4/historical/sports/{sport_key}/events"
    params = {"apiKey": api_key, "date": timestamp}
    response = requests.get(url, params=params)
    response.raise_for_status()
    return {
        "metaData": format_response_meta_data(response.headers),
        "responseContent": response.json()
    }


def fetch_event_odds(api_key, sport_key, event_id, regions, bookmakers, markets, odds_format, timestamp):
    """
    Fetch odds for a specific event from The Odds API.
    """
    bookmakers_param = f"bookmakers={bookmakers}" if bookmakers else f"regions={regions}"
    url = f"https://api.the-odds-api.com/v4/historical/sports/{sport_key}/events/{event_id}/odds"
    params = {
        "apiKey": api_key,
        bookmakers_param: bookmakers_param,
        "markets": markets,
        "oddsFormat": odds_format,
        "date": timestamp
    }
    response = requests.get(url, params=params)
    response.raise_for_status()
    return {
        "metaData": format_response_meta_data(response.headers),
        "responseContent": response.json()
    }


def format_event_output(response):
    """
    Restructure the JSON response into rows suitable for Excel output.
    """
    rows = []
    event = response["data"]
    for bookmaker in event.get("bookmakers", []):
        for market in bookmaker.get("markets", []):
            for outcome in market.get("outcomes", []):
                rows.append([
                    response["timestamp"], event["id"], event["commence_time"],
                    bookmaker["key"], event["home_team"], event["away_team"],
                    market["key"], market["last_update"], outcome["name"],
                    outcome.get("description"), outcome["price"], outcome.get("point")
                ])
    return rows


def format_response_meta_data(headers):
    """
    Extract metadata from response headers.
    """
    return [
        ["Requests Used", headers.get("x-requests-used")],
        ["Requests Remaining", headers.get("x-requests-remaining")]
    ]


def main():
    """
    Main function to query historical odds data and save it to an Excel spreadsheet.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # Write headers to the Excel file
    ws.append(HEADERS)

    # Parse date strings into datetime objects
    from_datetime = datetime.fromisoformat(FROM_DATE[:-1])
    to_datetime = datetime.fromisoformat(TO_DATE[:-1])
    current_datetime = to_datetime

    # Initialize variables
    output_row = 2
    data = None

    while current_datetime > from_datetime:
        formatted_date = current_datetime.isoformat() + "Z"
        print(f"Fetching events for {formatted_date}")

        # Fetch events
        events_response = fetch_events(API_KEY, SPORT_KEY, formatted_date)
        for event in events_response["responseContent"].get("data", []):
            event_id = event["id"]
            # Fetch odds for each event
            odds_data = fetch_event_odds(API_KEY, SPORT_KEY, event_id, REGIONS, BOOKMAKERS, MARKETS, ODDS_FORMAT,
                                         formatted_date)

            # Output metadata
            meta_data = odds_data["metaData"]
            for i, meta_row in enumerate(meta_data):
                ws.cell(row=i + 1, column=1, value=meta_row[0])
                ws.cell(row=i + 1, column=2, value=meta_row[1])

            # Output event odds data
            formatted_response = format_event_output(odds_data["responseContent"])
            for row in formatted_response:
                ws.append(row)

            output_row += len(formatted_response)

        # Update current_datetime to fetch earlier data
        previous_timestamp = events_response["responseContent"].get("previous_timestamp")
        if not previous_timestamp:
            print("No earlier historical data available.")
            break
        current_datetime = min(
            current_datetime - timedelta(minutes=INTERVAL_MINS),
            datetime.fromisoformat(previous_timestamp[:-1])
        )

    # Save the workbook to a file
    wb.save(SPREADSHEET_FILE)
    print(f"Data saved to {SPREADSHEET_FILE}")


if __name__ == "__main__":
    main()
